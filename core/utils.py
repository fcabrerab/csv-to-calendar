from openpyxl import load_workbook


def remove_all_pages_but_first(excel_file_name):
    """
    Remove all pages 2-n in the given Excel file.
    """
    wb = load_workbook(excel_file_name)
    if len(wb.sheetnames) > 1:
        wb.remove(wb.worksheets[1:])
    wb.save(excel_file_name)


def xlsx_to_csv(excel_file_name, csv_file_name):
    """
    Convert the given Excel file to CSV.
    """
    wb = load_workbook(excel_file_name)
    ws = wb.active
    with open(csv_file_name, "w") as csv_file:
        for row in ws.rows:
            for cell in row:
                csv_file.write(str(cell.value) + ",")
            csv_file.write("\n")
    wb.close()


def copy_file(source_file_name, destination_file_name):
    """
    Copy the given file.
    """
    with open(source_file_name, "rb") as source_file:
        with open(destination_file_name, "wb") as destination_file:
            destination_file.write(source_file.read())


def remove_file(file_name):
    """
    Remove the given file.
    """
    import os

    os.remove(file_name)
